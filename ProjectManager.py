import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from colorama import Fore, Style, init

init(autoreset=True)  # Colorama init para terminal

# CONFIGURA TU RUTA AQUÍ
ruta_proyecto = "C:\Proyectos\Demo-Huawei-DeepSeek"

estructura = defaultdict(lambda: {
    "llamado_por_bat": [],
    "usa_json": [],
    "usa_html": [],
    "usa_py": [],
})

# Analiza los archivos .bat
for root, _, files in os.walk(ruta_proyecto):
    for archivo in files:
        if archivo.endswith(".bat"):
            ruta_bat = os.path.join(root, archivo)
            with open(ruta_bat, "r", encoding="utf-8", errors="ignore") as f:
                contenido = f.read()
                llamados = re.findall(r"python\s+(.*?\.py)", contenido)
                for llamado in llamados:
                    estructura[llamado]["llamado_por_bat"].append(os.path.relpath(ruta_bat, ruta_proyecto))

# Analiza los archivos .py
for root, _, files in os.walk(ruta_proyecto):
    for archivo in files:
        if archivo.endswith(".py"):
            ruta_py = os.path.join(root, archivo)
            with open(ruta_py, "r", encoding="utf-8", errors="ignore") as f:
                contenido = f.read()
                estructura[archivo]["usa_json"] = re.findall(r"['\"](.*?\.json)['\"]", contenido)
                estructura[archivo]["usa_html"] = re.findall(r"['\"](.*?\.html)['\"]", contenido)
                estructura[archivo]["usa_py"] += re.findall(r"(?:from|import)\s+([a-zA-Z_][\w]*)", contenido)

# Crear Excel
wb = Workbook()

# Colores y estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")  # azul
gray_fill = PatternFill("solid", fgColor="D9D9D9")

def aplicar_estilos(ws):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(15, max_length + 2)

# --- Hoja 1: Resumen general ---
ws1 = wb.active
ws1.title = "Resumen General"
ws1.append(["Archivo PY", "Llamado por BAT", "Usa JSON", "Usa HTML", "Importa módulos"])

for script, data in estructura.items():
    llamado_por_bat = ", ".join(data["llamado_por_bat"])
    usa_json = ", ".join(data["usa_json"])
    usa_html = ", ".join(data["usa_html"])
    imports = ", ".join(sorted(set(data["usa_py"])))
    ws1.append([script, llamado_por_bat, usa_json, usa_html, imports])

aplicar_estilos(ws1)

# --- Hoja 2: BAT → PY ---
ws2 = wb.create_sheet("BAT → PY")
ws2.append(["Archivo BAT", "Ejecuta Script PY"])

for script, data in estructura.items():
    for bat in data["llamado_por_bat"]:
        ws2.append([bat, script])

aplicar_estilos(ws2)

# --- Hoja 3: PY → Recursos ---
ws3 = wb.create_sheet("PY → Recursos")
ws3.append(["Archivo PY", "Tipo", "Archivo Referenciado"])

for script, data in estructura.items():
    for json_file in data["usa_json"]:
        ws3.append([script, "JSON", json_file])
    for html_file in data["usa_html"]:
        ws3.append([script, "HTML", html_file])
    for imp in set(data["usa_py"]):
        ws3.append([script, "Import", imp])

aplicar_estilos(ws3)

# Guardar archivo
excel_path = os.path.join(ruta_proyecto, "estructura_proyecto_detallada.xlsx")
wb.save(excel_path)

print(f"{Fore.GREEN}\n✅ ¡Excel profesional generado con éxito!")
print(f"{Fore.CYAN}📄 Ruta: {excel_path}\n")
